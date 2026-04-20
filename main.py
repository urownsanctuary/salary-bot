import os
import uuid
import hashlib
from io import BytesIO
from html import escape
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, Depends, HTTPException, Form, UploadFile, File, Cookie
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.db import SessionLocal, engine
from app.services import (
    get_active_period,
    login_user,
    get_merchants_columns,
    normalize_point_code,
    point_has_any_supply_in_month,
    get_supply_boxes_map,
    get_visits_for_month,
    get_merchant_by_fio,
    toggle_day_visit,
    toggle_inventory_visit,
    compute_point_total,
    compute_overall_total,
    days_in_month,
    weekday_of,
    month_title,
    get_monthly_submission,
    upsert_monthly_submission_draft,
    submit_monthly_submission,
    reopen_monthly_submission,
    get_admin_report_rows,
    get_admin_payroll_rows,
    get_intersections_rows,
    get_all_tu_values,
    import_supplies_xlsx,
    import_rates_xlsx,
    import_merchants_xlsx,
    clear_month_data,
    clear_merchants_by_tu,
    get_point_adjustment,
    upsert_point_adjustment,
)

app = FastAPI()

app.mount("/static", StaticFiles(directory="app/static"), name="static")

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")

ADMIN_LOGIN = os.getenv("ADMIN_LOGIN", "")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "")
SECRET_SALT = os.getenv("SECRET_SALT", "")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_admin_cookie_value() -> str:
    raw = f"{ADMIN_LOGIN}:{ADMIN_PASSWORD}:{SECRET_SALT}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def is_admin_authenticated(admin_auth: Optional[str]) -> bool:
    if not ADMIN_LOGIN or not ADMIN_PASSWORD or not SECRET_SALT:
        return False
    return admin_auth == get_admin_cookie_value()


def style_sheet(ws):
    green_fill = PatternFill("solid", fgColor="E8F5E9")
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 35)
    ws.freeze_panes = "A2"


def build_excel_response(wb: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def base_css():
    return """
    <style>
        @font-face {
            font-family: 'Villula';
            src: url('/static/fonts/villula-regular.ttf') format('truetype');
            font-weight: normal;
            font-style: normal;
        }

        :root {
            --bg: #F6F8F7;
            --card: #FFFFFF;
            --text: #1F2937;
            --muted: #6B7280;
            --line: #D1D5DB;
            --green: #2E7D32;
            --green-dark: #27682A;
            --soft: #EEF4EF;
            --soft-2: #F3F7F3;
            --error: #B91C1C;
            --ok: #166534;
            --shadow: 0 12px 32px rgba(0, 0, 0, 0.08);
        }

        * { box-sizing: border-box; }

        body {
            margin: 0;
            background: var(--bg);
            color: var(--text);
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
            min-height: 100vh;
            padding: 20px;
        }

        .page {
            max-width: 1580px;
            margin: 0 auto;
            min-height: calc(100vh - 40px);
            display: flex;
            align-items: center;
            justify-content: center;
        }

        .card {
            width: 100%;
            max-width: 430px;
            background: var(--card);
            border-radius: 24px;
            padding: 32px 28px;
            box-shadow: var(--shadow);
        }

        .card-wide {
            width: 100%;
            max-width: 1580px;
            background: var(--card);
            border-radius: 24px;
            padding: 22px 20px 28px;
            box-shadow: var(--shadow);
        }

        .brand {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 28px;
            line-height: 1;
            color: var(--green);
            margin-bottom: 8px;
        }

        h1 {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 34px;
            line-height: 1.05;
            margin: 0 0 8px 0;
            color: var(--text);
        }

        .subtitle {
            color: var(--muted);
            font-size: 15px;
            line-height: 1.45;
            margin-bottom: 18px;
        }

        label {
            display: block;
            margin: 14px 0 6px;
            font-size: 14px;
            font-weight: 700;
            color: var(--text);
        }

        input, textarea, select {
            width: 100%;
            padding: 14px 16px;
            border: 1px solid var(--line);
            border-radius: 14px;
            font-size: 16px;
            background: #fff;
            font-family: inherit;
        }

        textarea {
            resize: vertical;
            min-height: 92px;
        }

        input:focus, textarea:focus, select:focus {
            outline: none;
            border-color: var(--green);
            box-shadow: 0 0 0 3px rgba(46, 125, 50, 0.10);
        }

        .btn {
            display: inline-block;
            width: 100%;
            margin-top: 16px;
            padding: 15px 16px;
            border: none;
            border-radius: 14px;
            background: var(--green);
            color: #fff;
            font-size: 16px;
            font-weight: 800;
            text-align: center;
            text-decoration: none;
            cursor: pointer;
        }

        .btn:hover { background: var(--green-dark); }

        .btn-secondary {
            background: var(--soft);
            color: var(--text);
        }

        .btn-secondary:hover { background: #e3ece3; }

        .btn-danger {
            background: #B91C1C;
            color: #fff;
        }

        .btn-danger:hover {
            background: #991B1B;
        }

        .btn-small {
            margin-top: 12px;
            padding: 12px 14px;
            font-size: 15px;
        }

        .btn-inline {
            width: auto;
            margin-top: 0;
            padding: 12px 16px;
            font-size: 14px;
        }

        .footer {
            margin-top: 18px;
            color: #9CA3AF;
            font-size: 12px;
            text-align: center;
        }

        .back {
            display: inline-block;
            margin-top: 18px;
            color: var(--green);
            text-decoration: none;
            font-weight: 800;
        }

        .hint {
            margin-top: 16px;
            padding: 12px 14px;
            border-radius: 14px;
            background: var(--soft);
            color: var(--muted);
            font-size: 13px;
            line-height: 1.4;
        }

        .error-box {
            margin-top: 16px;
            background: #FEF2F2;
            color: var(--error);
            border-radius: 14px;
            padding: 14px;
            line-height: 1.45;
            font-weight: 700;
        }

        .success-box {
            margin-top: 16px;
            background: #ECFDF3;
            color: var(--ok);
            border-radius: 14px;
            padding: 14px;
            line-height: 1.45;
            font-weight: 700;
        }

        .calendar-head {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 14px;
            flex-wrap: wrap;
        }

        .calendar-month {
            font-family: 'Villula', -apple-system, sans-serif;
            font-size: 28px;
            line-height: 1;
        }

        .calendar-meta {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
        }

        .mini-pill {
            background: var(--soft-2);
            border-radius: 999px;
            padding: 8px 12px;
            font-size: 13px;
            color: var(--text);
            font-weight: 700;
        }

        .sum-strip {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
            margin-bottom: 16px;
        }

        .sum-card {
            background: #F7FBF8;
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            padding: 14px 16px;
        }

        .sum-title {
            color: var(--muted);
            font-size: 13px;
            margin-bottom: 6px;
        }

        .sum-value {
            font-size: 22px;
            font-weight: 900;
        }

        .details-grid {
            display: grid;
            grid-template-columns: repeat(2, 1fr);
            gap: 12px;
            margin-bottom: 18px;
        }

        .detail-card {
            background: #FAFCFA;
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            padding: 14px 16px;
        }

        .detail-title {
            color: var(--muted);
            font-size: 13px;
            margin-bottom: 8px;
        }

        .detail-line {
            font-size: 15px;
            font-weight: 700;
            line-height: 1.5;
        }

        .calendar-wrap { margin-top: 4px; }

        .weekdays, .calendar-grid {
            display: grid;
            grid-template-columns: repeat(7, 1fr);
            gap: 10px;
        }

        .weekdays { margin-bottom: 10px; }

        .weekday {
            text-align: center;
            font-size: 13px;
            color: var(--muted);
            font-weight: 700;
            padding: 6px 0;
        }

        .day, .day-empty {
            min-height: 90px;
            border-radius: 18px;
            padding: 10px;
        }

        .day {
            background: #F8FAF8;
            border: 1px solid #E5E7EB;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            text-decoration: none;
            color: inherit;
            cursor: pointer;
        }

        .day:hover {
            border-color: var(--green);
            box-shadow: 0 0 0 2px rgba(46, 125, 50, 0.06);
        }

        .day-empty { background: transparent; }

        .day-disabled {
            opacity: 0.65;
            cursor: default;
            pointer-events: none;
        }

        .day-number {
            font-size: 18px;
            font-weight: 800;
        }

        .day-badges {
            display: flex;
            flex-wrap: wrap;
            gap: 6px;
            margin-top: 10px;
        }

        .badge {
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 4px 8px;
            border-radius: 999px;
            font-size: 11px;
            font-weight: 800;
            line-height: 1;
            min-width: 22px;
            height: 22px;
        }

        .badge-supply {
            background: #2E7D32;
            color: #fff;
            border-radius: 6px;
        }

        .badge-day {
            background: #DBEAFE;
            color: #1D4ED8;
        }

        .badge-inv {
            background: #FCE7F3;
            color: #BE185D;
        }

        .legend {
            margin-top: 18px;
            display: flex;
            flex-wrap: wrap;
            gap: 10px;
        }

        .legend-item {
            background: var(--soft);
            border-radius: 999px;
            padding: 8px 12px;
            font-size: 13px;
            color: var(--text);
            font-weight: 700;
        }

        .calendar-note {
            margin-top: 14px;
            color: var(--muted);
            line-height: 1.5;
            font-size: 14px;
        }

        details.point-detail {
            margin-top: 10px;
            border: 1px solid #E5E7EB;
            border-radius: 14px;
            background: #FAFCFA;
            padding: 10px 14px;
        }

        details.point-detail summary {
            cursor: pointer;
            font-weight: 800;
            list-style: none;
        }

        details.point-detail summary::-webkit-details-marker {
            display: none;
        }

        .summary-content {
            margin-top: 10px;
            color: var(--text);
            line-height: 1.6;
        }

        .filter-grid {
            display: grid;
            grid-template-columns: 140px 140px 240px 220px;
            gap: 12px;
            margin-bottom: 16px;
            align-items: end;
        }

        .table-wrap {
            border: 1px solid #E5E7EB;
            border-radius: 16px;
            background: #fff;
            overflow: visible;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            table-layout: fixed;
        }

        th, td {
            padding: 10px 8px;
            border-bottom: 1px solid #E5E7EB;
            text-align: left;
            vertical-align: top;
            font-size: 12px;
            word-break: break-word;
        }

        th {
            background: #F7FBF8;
            font-weight: 800;
        }

        .admin-actions {
            display: flex;
            gap: 12px;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
            flex-wrap: wrap;
        }

        .admin-export-buttons {
            display: flex;
            gap: 10px;
            flex-wrap: wrap;
            margin-top: 14px;
        }

        .data-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
        }

        @media (max-width: 960px) {
            .page { align-items: flex-start; }
            .card-wide { padding: 18px 14px 24px; }
            .sum-strip, .details-grid, .filter-grid, .data-grid { grid-template-columns: 1fr; }
            .weekdays, .calendar-grid { gap: 8px; }
            .day, .day-empty {
                min-height: 80px;
                border-radius: 14px;
                padding: 8px;
            }
            .day-number { font-size: 16px; }
            h1 { font-size: 30px; }
            .brand { font-size: 24px; }
            .calendar-month { font-size: 24px; }
            .table-wrap { overflow-x: auto; }
            table { min-width: 1200px; }
        }
    </style>
    """


@app.get("/")
def root():
    return RedirectResponse(url="/login-page")


@app.get("/db-check")
def db_check():
    with engine.connect() as conn:
        conn.execute(text("SELECT 1"))
    return {"status": "ok", "db": "connected"}


@app.get("/active-period")
def active_period():
    return get_active_period()


@app.get("/debug/merchants-columns")
def merchants_columns(db: Session = Depends(get_db)):
    cols = get_merchants_columns(db)
    return {"table": "merchants", "columns": cols}


@app.post("/login")
def login_api(fio: str, last4: str, db: Session = Depends(get_db)):
    user = login_user(db, fio, last4)

    if not user:
        raise HTTPException(status_code=401, detail="Неверные данные")

    return {"status": "ok", "active_period": get_active_period(), "user": user}


@app.get("/login-page", response_class=HTMLResponse)
def login_page():
    period = get_active_period()
    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>ВкусВилл</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Сверки мерчендайзеров</h1>
            <div class="subtitle">Введите ФИО и последние 4 цифры телефона</div>

            <form method="post" action="/login-page">
                <label for="fio">ФИО</label>
                <input id="fio" name="fio" type="text" placeholder="Иванов Иван Иванович" required />

                <label for="last4">Последние 4 цифры телефона</label>
                <input id="last4" name="last4" type="text" inputmode="numeric" maxlength="4" placeholder="1234" required />

                <button class="btn" type="submit">Войти</button>
            </form>

            <div class="hint">Сейчас открыт период за {month_title(period["year"], period["month"])}.</div>

            <div class="footer">Веб-версия сверок мерчендайзеров</div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/login-page", response_class=HTMLResponse)
def login_submit(
    fio: str = Form(...),
    last4: str = Form(...),
    db: Session = Depends(get_db)
):
    user = login_user(db, fio, last4)

    if not user:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Ошибка входа</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>Ошибка входа</h1>
            <div class="error-box">Неверные данные. Проверьте ФИО и последние 4 цифры телефона.</div>
            <a class="back" href="/login-page">← Попробовать снова</a>
        </div>
    </div>
</body>
</html>
"""

    return RedirectResponse(url=f"/menu-page?fio={user['fio']}", status_code=303)


@app.get("/menu-page", response_class=HTMLResponse)
def menu_page(fio: str = "", db: Session = Depends(get_db)):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    overall = {"total": 0}
    if merchant:
        overall = compute_overall_total(db, merchant["id"], period["year"], period["month"])

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Главное меню</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Главное меню</h1>
            <div class="subtitle">{escape(fio)}</div>
            <div class="hint">Сейчас открыт период за {month_title(period["year"], period["month"])}.</div>

            <div class="sum-card" style="margin-top: 18px;">
                <div class="sum-title">Общая сумма за месяц</div>
                <div class="sum-value">{overall["total"]} ₽</div>
            </div>

            <a class="btn" href="/point-page?fio={escape(fio)}">Заполнить сверку</a>
            <a class="btn btn-secondary" href="/summary-page?fio={escape(fio)}">Моя сумма</a>
            <a class="btn btn-secondary" href="/monthly-submit-page?fio={escape(fio)}">Отправить сверку за месяц</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/point-page", response_class=HTMLResponse)
def point_page(fio: str = ""):
    period = get_active_period()

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Выбор точки</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Выбор точки</h1>
            <div class="subtitle">{escape(fio)}</div>

            <div class="hint" style="margin-top: 0; margin-bottom: 18px;">
                Сверка заполняется за {month_title(period["year"], period["month"])}.
            </div>

            <form method="post" action="/point-page">
                <input type="hidden" name="fio" value="{escape(fio)}" />

                <label for="point_code">Номер точки</label>
                <input id="point_code" name="point_code" type="text" placeholder="2674" required />

                <button class="btn" type="submit">Продолжить</button>
            </form>

            <a class="back" href="/menu-page?fio={escape(fio)}">← Назад</a>
        </div>
    </div>
</body>
</html>
"""


@app.post("/point-page", response_class=HTMLResponse)
def point_submit(
    fio: str = Form(...),
    point_code: str = Form(...),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    point_code = normalize_point_code(point_code)

    if not point_code or len(point_code) < 3:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Ошибка</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>Ошибка</h1>
            <div class="error-box">Номер точки слишком короткий.</div>
            <a class="back" href="/point-page?fio={escape(fio)}">← Назад</a>
        </div>
    </div>
</body>
</html>
"""

    has_supply = point_has_any_supply_in_month(db, point_code, period["year"], period["month"])

    if not has_supply:
        return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Точка не найдена</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <h1>Точка не найдена</h1>
            <div class="error-box">
                В периоде {month_title(period["year"], period["month"])} по точке {escape(point_code)} нет поставок.
                <br><br>
                Проверьте номер точки или обратитесь к управляющему.
            </div>
            <a class="back" href="/point-page?fio={escape(fio)}">← Попробовать снова</a>
        </div>
    </div>
</body>
</html>
"""

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


def build_day_href(fio: str, point_code: str, y: int, m: int, day: int, is_submitted: bool) -> str:
    if is_submitted:
        return "#"
    wd = weekday_of(y, m, day)
    if wd in (4, 5):
        return f"/day-action-page?fio={escape(fio)}&point_code={escape(point_code)}&day={day}"
    return f"/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}"


def build_calendar_html(
    fio: str,
    point_code: str,
    y: int,
    m: int,
    boxes_map: dict[int, int],
    visits: dict[int, set[str]],
    is_submitted: bool
) -> str:
    dim = days_in_month(y, m)
    first_wd = weekday_of(y, m, 1)
    weekdays = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

    html = '<div class="weekdays">'
    for wd in weekdays:
        html += f'<div class="weekday">{wd}</div>'
    html += '</div>'

    html += '<div class="calendar-grid">'

    for _ in range(first_wd):
        html += '<div class="day-empty"></div>'

    for day in range(1, dim + 1):
        boxes = boxes_map.get(day, 0)
        day_visits = visits.get(day, set())

        badges = ""
        if boxes > 0:
            badges += '<span class="badge badge-supply">П</span>'
        if "DAY" in day_visits:
            badges += '<span class="badge badge-day">В</span>'
        if "FULL_INVENT" in day_visits:
            badges += '<span class="badge badge-inv">И</span>'

        href = build_day_href(fio, point_code, y, m, day, is_submitted)
        cls = "day day-disabled" if is_submitted else "day"

        html += f"""
        <a class="{cls}" href="{href}">
            <div class="day-number">{day}</div>
            <div class="day-badges">{badges}</div>
        </a>
        """

    html += '</div>'
    return html


@app.get("/calendar-page", response_class=HTMLResponse)
def calendar_page(
    fio: str,
    point_code: str,
    saved: str = "",
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    point_code = normalize_point_code(point_code)

    overall = compute_overall_total(db, merchant["id"], y, m)
    monthly_submitted = overall["submission_status"] == "submitted"

    boxes_map = get_supply_boxes_map(db, point_code, y, m)
    visits = get_visits_for_month(db, merchant["id"], point_code, y, m)
    point_total = compute_point_total(db, merchant["id"], point_code, y, m)
    point_adj = get_point_adjustment(db, merchant["id"], point_code, y, m) or {}

    calendar_html = build_calendar_html(
        fio=fio,
        point_code=point_code,
        y=y,
        m=m,
        boxes_map=boxes_map,
        visits=visits,
        is_submitted=monthly_submitted
    )

    info_box = ""
    if saved == "1":
        info_box = "<div class=\"success-box\">Данные по точке сохранены.</div>"

    point_receipt_link = ""
    if point_total["reimb_receipt"]:
        point_receipt_link = f"<div class='hint' style='margin-top:10px'>Чек по точке: <a href='/{point_total['reimb_receipt']}' target='_blank'>открыть файл</a></div>"

    point_form = ""
    if not monthly_submitted:
        point_form = f"""
            <form method="post" action="/save-point-adjustment" enctype="multipart/form-data" class="detail-card" style="margin-top:18px;">
                <input type="hidden" name="fio" value="{escape(fio)}" />
                <input type="hidden" name="point_code" value="{escape(point_code)}" />

                <div class="detail-title">Примечание по точке</div>
                <label for="note_amount">Сумма, ₽</label>
                <input id="note_amount" name="note_amount" type="number" min="0" value="{point_total['note_amount']}" placeholder="Например: 1500" />

                <label for="note_comment">Комментарий</label>
                <input id="note_comment" name="note_comment" type="text" value="{escape(point_total['note_comment'])}" placeholder="Например: Закрытие точки" />

                <div class="detail-title" style="margin-top:18px;">Возмещение по точке</div>
                <label for="reimb_amount">Сумма, ₽</label>
                <input id="reimb_amount" name="reimb_amount" type="number" min="0" value="{point_total['reimb_amount']}" placeholder="Например: 150" />

                <label for="reimb_comment">Комментарий</label>
                <input id="reimb_comment" name="reimb_comment" type="text" value="{escape(point_total['reimb_comment'])}" placeholder="Например: Покупка пакетов" />

                <label for="reimb_receipt">Чек</label>
                <input id="reimb_receipt" name="reimb_receipt" type="file" accept=".jpg,.jpeg,.png,.pdf,.webp" />
                <div class="hint" style="margin-top:10px;">Если указано возмещение, чек обязателен.</div>
                {point_receipt_link}

                <button class="btn btn-secondary" type="submit">Сохранить данные по точке</button>
            </form>
        """

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Календарь</title>
    {base_css()}
    <script>
    document.addEventListener('DOMContentLoaded', function() {{
        const savedY = sessionStorage.getItem('calendarScrollY');
        if (savedY) {{
            window.scrollTo(0, parseInt(savedY, 10));
            sessionStorage.removeItem('calendarScrollY');
        }}
        document.querySelectorAll('.day').forEach(el => {{
            el.addEventListener('click', function() {{
                sessionStorage.setItem('calendarScrollY', String(window.scrollY));
            }});
        }});
    }});
    </script>
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="calendar-head">
                <div>
                    <div class="brand">ВкусВилл</div>
                    <div class="calendar-month">{month_title(y, m)}</div>
                </div>

                <div class="calendar-meta">
                    <div class="mini-pill">Точка: {escape(point_code)}</div>
                    <div class="mini-pill">{escape(fio)}</div>
                    <div class="mini-pill">КМ: {"Да" if point_total["coffee_enabled"] else "Нет"}</div>
                    <div class="mini-pill">Месячная сверка: {"Отправлена" if monthly_submitted else "Черновик"}</div>
                </div>
            </div>

            {info_box}

            <div class="sum-strip">
                <div class="sum-card">
                    <div class="sum-title">Сумма по точке</div>
                    <div class="sum-value">{point_total["total"]} ₽</div>
                </div>

                <div class="sum-card">
                    <div class="sum-title">Общая сумма за месяц</div>
                    <div class="sum-value">{overall["total"]} ₽</div>
                </div>
            </div>

            <div class="details-grid">
                <div class="detail-card">
                    <div class="detail-title">Выходы с поставкой</div>
                    <div class="detail-line">{point_total["cnt_supply"]} × {point_total["rate_supply"]} ₽ = {point_total["sum_supply"]} ₽</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Выходы без поставки</div>
                    <div class="detail-line">{point_total["cnt_no_supply"]} × {point_total["rate_no_supply"]} ₽ = {point_total["sum_no_supply"]} ₽</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Полные инвенты</div>
                    <div class="detail-line">{point_total["cnt_full_inv"]} × {point_total["rate_inventory"]} ₽ = {point_total["sum_inventory"]} ₽</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Кофемашина</div>
                    <div class="detail-line">{point_total["coffee_cnt"]} × {point_total["coffee_rate"]} ₽ = {point_total["coffee_sum"]} ₽</div>
                </div>
            </div>

            <div class="details-grid">
                <div class="detail-card">
                    <div class="detail-title">Примечание по точке</div>
                    <div class="detail-line">{point_total["note_amount"]} ₽</div>
                    <div class="calendar-note">{escape(point_total["note_comment"]) if point_total["note_comment"] else "—"}</div>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Возмещение по точке</div>
                    <div class="detail-line">{point_total["reimb_amount"]} ₽</div>
                    <div class="calendar-note">{escape(point_total["reimb_comment"]) if point_total["reimb_comment"] else "—"}</div>
                </div>
            </div>

            <div class="calendar-wrap">
                {calendar_html}
            </div>

            <div class="legend">
                <div class="legend-item">П — была поставка</div>
                <div class="legend-item">В — отмечен выход</div>
                <div class="legend-item">И — полный инвент</div>
            </div>

            <div class="calendar-note">
                В обычные дни нажатие по дню сразу ставит или убирает выход.
                В пятницу и субботу открывается выбор: выход или полный инвент.
            </div>

            <div class="calendar-note">
                Поставки до 5 коробок не оплачиваются.
            </div>

            {point_form}

            <div class="admin-export-buttons" style="margin-top:18px;">
                <a class="btn btn-secondary btn-inline" href="/point-page?fio={escape(fio)}">Следующая точка</a>
                <a class="btn btn-secondary btn-inline" href="/summary-page?fio={escape(fio)}">Моя сумма</a>
                <a class="btn btn-secondary btn-inline" href="/monthly-submit-page?fio={escape(fio)}">Отправить сверку за месяц</a>
            </div>

            <a class="back" href="/menu-page?fio={escape(fio)}">← На главный экран</a>
        </div>
    </div>
</body>
</html>
"""


@app.post("/save-point-adjustment")
async def save_point_adjustment(
    fio: str = Form(...),
    point_code: str = Form(...),
    note_amount: int = Form(0),
    note_comment: str = Form(""),
    reimb_amount: int = Form(0),
    reimb_comment: str = Form(""),
    reimb_receipt: UploadFile | None = File(None),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    receipt_path = None
    if reimb_receipt and reimb_receipt.filename:
        ext = Path(reimb_receipt.filename).suffix.lower()
        filename = f"{uuid.uuid4().hex}{ext}"
        filepath = UPLOAD_DIR / filename
        content = await reimb_receipt.read()
        filepath.write_bytes(content)
        receipt_path = f"uploads/{filename}"

    upsert_point_adjustment(
        db=db,
        merchant_id=merchant["id"],
        point_code=normalize_point_code(point_code),
        y=period["year"],
        m=period["month"],
        note_amount=max(0, int(note_amount or 0)),
        note_comment=note_comment or "",
        reimb_amount=max(0, int(reimb_amount or 0)),
        reimb_comment=reimb_comment or "",
        reimb_receipt=receipt_path,
    )

    return RedirectResponse(
        url=f"/calendar-page?fio={escape(fio)}&point_code={escape(normalize_point_code(point_code))}&saved=1",
        status_code=303
    )



@app.get("/monthly-submit-page", response_class=HTMLResponse)
def monthly_submit_page(
    fio: str,
    submitted: str = "",
    reopened: str = "",
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    monthly_submitted = overall["submission_status"] == "submitted"

    info_box = ""
    if submitted == "1":
        info_box += "<div class='success-box'>Месячная сверка отправлена.</div>"
    if reopened == "1":
        info_box += "<div class='success-box'>Месячная сверка разблокирована для редактирования.</div>"

    points_html = ""
    if overall["per_point_details"]:
        for point_code, d in overall["per_point_details"].items():
            points_html += f"""
            <details class="point-detail">
                <summary>{escape(point_code)} — {d["total"]} ₽</summary>
                <div class="summary-content">
                    <div>С поставкой: {d["cnt_supply"]} × {d["rate_supply"]} ₽ = {d["sum_supply"]} ₽</div>
                    <div>Без поставки: {d["cnt_no_supply"]} × {d["rate_no_supply"]} ₽ = {d["sum_no_supply"]} ₽</div>
                    <div>Полный инвент: {d["cnt_full_inv"]} × {d["rate_inventory"]} ₽ = {d["sum_inventory"]} ₽</div>
                    <div>Кофемашина: {d["coffee_cnt"]} × {d["coffee_rate"]} ₽ = {d["coffee_sum"]} ₽</div>
                    <div>Примечание по точке: {d["note_amount"]} ₽ — {escape(d["note_comment"]) if d["note_comment"] else "—"}</div>
                    <div>Возмещение по точке: {d["reimb_amount"]} ₽ — {escape(d["reimb_comment"]) if d["reimb_comment"] else "—"}</div>
                    <div>Чек по возмещению: {f"<a href='/{d['reimb_receipt']}' target='_blank'>открыть</a>" if d["reimb_receipt"] else "—"}</div>
                </div>
            </details>
            """
    else:
        points_html = "<div class='hint'>В этом месяце пока нет отмеченных точек.</div>"

    action_block = ""
    if monthly_submitted:
        action_block = f"""
        <div class="detail-card" style="margin-top:18px;">
            <div class="detail-title">Статус</div>
            <div class="detail-line">Сверка за месяц отправлена</div>
            <a class="btn btn-secondary" href="/reopen-monthly-submission?fio={escape(fio)}">Редактировать сверку</a>
        </div>
        """
    else:
        action_block = f"""
        <form method="post" action="/submit-monthly-submission">
            <input type="hidden" name="fio" value="{escape(fio)}" />
            <button class="btn" type="submit">Отправить сверку за месяц</button>
        </form>
        """

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Отправка месячной сверки</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="brand">ВкусВилл</div>
            <h1>Отправить сверку за месяц</h1>
            <div class="subtitle">{escape(fio)} · {month_title(y, m)}</div>

            {info_box}

            <div class="sum-strip">
                <div class="sum-card">
                    <div class="sum-title">Сумма по точкам</div>
                    <div class="sum-value">{sum(overall["per_point"].values())} ₽</div>
                </div>

                <div class="sum-card">
                    <div class="sum-title">Итог за месяц</div>
                    <div class="sum-value">{overall["total"]} ₽</div>
                </div>
            </div>

            <div class="hint">
                На этой странице больше нет полей примечания и возмещения за месяц.
                Они заполняются отдельно внутри каждой точки.
            </div>

            {points_html}

            {action_block}

            <div class="admin-export-buttons" style="margin-top:18px;">
                <a class="btn btn-secondary btn-inline" href="/point-page?fio={escape(fio)}">Перейти к другой точке</a>
                <a class="btn btn-secondary btn-inline" href="/summary-page?fio={escape(fio)}">Моя сумма</a>
                <a class="btn btn-secondary btn-inline" href="/menu-page?fio={escape(fio)}">Главный экран</a>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/submit-monthly-submission")
async def submit_monthly_submission_route(
    fio: str = Form(...),
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    submit_monthly_submission(db, merchant["id"], period["year"], period["month"])

    return RedirectResponse(
        url=f"/monthly-submit-page?fio={escape(fio)}&submitted=1",
        status_code=303
    )


@app.get("/reopen-monthly-submission")
def reopen_monthly_submission_route(
    fio: str,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    reopen_monthly_submission(db, merchant["id"], period["year"], period["month"])

    return RedirectResponse(
        url=f"/monthly-submit-page?fio={escape(fio)}&reopened=1",
        status_code=303
    )

@app.get("/day-action-page", response_class=HTMLResponse)
def day_action_page(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if day < 1 or day > days_in_month(y, m):
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    visits = get_visits_for_month(db, merchant["id"], point_code, y, m)
    day_visits = visits.get(day, set())

    wd = weekday_of(y, m, day)
    is_fri_or_sat = wd in (4, 5)

    day_btn_text = "Убрать выход" if "DAY" in day_visits else "Добавить выход"
    inv_btn_text = "Убрать полный инвент" if "FULL_INVENT" in day_visits else "Добавить полный инвент"

    if not is_fri_or_sat:
        return RedirectResponse(url=f"/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}", status_code=303)

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Действие по дню</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Выбор действия</h1>
            <div class="subtitle">
                Точка: {escape(point_code)}<br>
                Дата: {day:02d}.{m:02d}.{y}
            </div>

            <a class="btn btn-small" href="/toggle-day?fio={escape(fio)}&point_code={escape(point_code)}&day={day}">
                {day_btn_text}
            </a>

            <a class="btn btn-secondary btn-small" href="/toggle-inventory?fio={escape(fio)}&point_code={escape(point_code)}&day={day}">
                {inv_btn_text}
            </a>

            <a class="back" href="/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}">← Назад к календарю</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/toggle-day")
def toggle_day(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if 1 <= day <= days_in_month(y, m):
        toggle_day_visit(db, merchant["id"], point_code, y, m, day)

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


@app.get("/toggle-inventory")
def toggle_inventory(
    fio: str,
    point_code: str,
    day: int,
    db: Session = Depends(get_db)
):
    period = get_active_period()
    y = period["year"]
    m = period["month"]

    merchant = get_merchant_by_fio(db, fio)
    if not merchant:
        return RedirectResponse(url="/login-page", status_code=303)

    overall = compute_overall_total(db, merchant["id"], y, m)
    if overall["submission_status"] == "submitted":
        return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)

    if 1 <= day <= days_in_month(y, m):
        wd = weekday_of(y, m, day)
        if wd in (4, 5):
            toggle_inventory_visit(db, merchant["id"], point_code, y, m, day)

    return RedirectResponse(url=f"/calendar-page?fio={escape(fio)}&point_code={escape(point_code)}", status_code=303)


@app.get("/summary-page", response_class=HTMLResponse)
def summary_page(fio: str = "", db: Session = Depends(get_db)):
    period = get_active_period()
    merchant = get_merchant_by_fio(db, fio)
    overall = {"total": 0, "per_point": {}, "per_point_details": {}}

    if merchant:
        overall = compute_overall_total(db, merchant["id"], period["year"], period["month"])

    details_html = ""
    if overall["per_point_details"]:
        for point_code, d in overall["per_point_details"].items():
            details_html += f"""
            <details class="point-detail">
                <summary>{escape(point_code)} — {d["total"]} ₽</summary>
                <div class="summary-content">
                    <div>С поставкой: {d["cnt_supply"]} × {d["rate_supply"]} ₽ = {d["sum_supply"]} ₽</div>
                    <div>Без поставки: {d["cnt_no_supply"]} × {d["rate_no_supply"]} ₽ = {d["sum_no_supply"]} ₽</div>
                    <div>Полный инвент: {d["cnt_full_inv"]} × {d["rate_inventory"]} ₽ = {d["sum_inventory"]} ₽</div>
                    <div>Кофемашина: {d["coffee_cnt"]} × {d["coffee_rate"]} ₽ = {d["coffee_sum"]} ₽</div>
                    <div><strong>Итого по точке: {d["total"]} ₽</strong></div>
                </div>
            </details>
            """
    else:
        details_html = "<div class='hint' style='margin-top:10px'>Пока нет отмеченных точек за этот месяц.</div>"


    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Моя сумма</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Моя сумма</h1>
            <div class="subtitle">{escape(fio)}</div>

            <div class="sum-card">
                <div class="sum-title">Общая сумма за месяц</div>
                <div class="sum-value">{overall["total"]} ₽</div>
            </div>

            {details_html}

            <div class="hint">Сейчас открыт период за {month_title(period["year"], period["month"])}.</div>

            <a class="back" href="/menu-page?fio={escape(fio)}">← Назад</a>
        </div>
    </div>
</body>
</html>
"""


@app.get("/admin-login", response_class=HTMLResponse)
def admin_login_page(error: str = ""):
    error_box = ""
    if error == "1":
        error_box = "<div class='error-box'>Неверный логин или пароль.</div>"

    env_box = ""
    if not ADMIN_LOGIN or not ADMIN_PASSWORD:
        env_box = "<div class='error-box'>В Render нужно задать ADMIN_LOGIN и ADMIN_PASSWORD.</div>"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Админ-вход</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card">
            <div class="brand">ВкусВилл</div>
            <h1>Админка</h1>
            <div class="subtitle">Вход в отчёт по сверкам</div>

            {env_box}
            {error_box}

            <form method="post" action="/admin-login">
                <label for="login">Логин</label>
                <input id="login" name="login" type="text" required />

                <label for="password">Пароль</label>
                <input id="password" name="password" type="password" required />

                <button class="btn" type="submit">Войти</button>
            </form>
        </div>
    </div>
</body>
</html>
"""


@app.post("/admin-login")
def admin_login_submit(login: str = Form(...), password: str = Form(...)):
    if not ADMIN_LOGIN or not ADMIN_PASSWORD:
        return RedirectResponse(url="/admin-login?error=1", status_code=303)

    if login != ADMIN_LOGIN or password != ADMIN_PASSWORD:
        return RedirectResponse(url="/admin-login?error=1", status_code=303)

    response = RedirectResponse(url="/admin-report", status_code=303)
    response.set_cookie(
        key="admin_auth",
        value=get_admin_cookie_value(),
        httponly=True,
        samesite="lax",
        secure=False,
        max_age=60 * 60 * 12,
    )
    return response


@app.get("/admin-logout")
def admin_logout():
    response = RedirectResponse(url="/admin-login", status_code=303)
    response.delete_cookie("admin_auth")
    return response


@app.get("/admin-report", response_class=HTMLResponse)
def admin_report(
    year: int | None = None,
    month: int | None = None,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    period = get_active_period()
    year = year or period["year"]
    month = month or period["month"]

    tu_filter = tu.strip() or None
    status_filter = status.strip() or None

    rows = get_admin_report_rows(db, year, month, tu_filter, status_filter)
    tu_values = get_all_tu_values(db)

    tu_options = "<option value=''>Все ТУ</option>"
    for item in tu_values:
        selected = "selected" if item == tu else ""
        tu_options += f"<option value='{escape(item)}' {selected}>{escape(item)}</option>"

    status_options = f"""
        <option value='' {'selected' if not status else ''}>Все статусы</option>
        <option value='не отправлено' {'selected' if status == 'не отправлено' else ''}>Не отправлено</option>
        <option value='draft' {'selected' if status == 'draft' else ''}>Черновик</option>
        <option value='submitted' {'selected' if status == 'submitted' else ''}>Отправлено</option>
    """

    rows_html = ""
    if rows:
        for r in rows:
            receipt_html = "—"
            if r["receipt_path"]:
                receipt_html = f"<a href='/{r['receipt_path']}' target='_blank'>Открыть</a>"

            rows_html += f"""
            <tr>
                <td>{escape(r["fio"])}</td>
                <td>{escape(r["tu"]) if r["tu"] else "—"}</td>
                <td>{escape(r["point_code"])}</td>
                <td>{month_title(year, month)}</td>
                <td>{r["cnt_supply"]} / {r["sum_supply"]} ₽</td>
                <td>{r["cnt_no_supply"]} / {r["sum_no_supply"]} ₽</td>
                <td>{r["cnt_full_inv"]} / {r["sum_inventory"]} ₽</td>
                <td>{r["coffee_cnt"]} × {r["coffee_rate"]} = {r["coffee_sum"]} ₽</td>
                <td>{r["note_amount"]} ₽<br>{escape(r["note_comment"]) if r["note_comment"] else "—"}</td>
                <td>{r["reimb_amount"]} ₽<br>{escape(r["reimb_comment"]) if r["reimb_comment"] else "—"}</td>
                <td><strong>{r["point_total"]} ₽</strong></td>
                <td>{escape(r["status"])}</td>
                <td>{f"<a href='/{r['reimb_receipt']}' target='_blank'>Открыть</a>" if r["reimb_receipt"] else "—"}</td>
                <td>{escape(r["comment"]) if r["comment"] else "—"}</td>
            </tr>
            """
    else:
        rows_html = """
        <tr>
            <td colspan="13">По выбранным фильтрам данных нет.</td>
        </tr>
        """

    month_options = ""
    for m in range(1, 13):
        selected = "selected" if m == month else ""
        month_options += f"<option value='{m}' {selected}>{m:02d}</option>"

    export_query = f"year={year}&month={month}&tu={escape(tu)}&status={escape(status)}"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Админ-отчёт</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="admin-actions">
                <div>
                    <div class="brand">ВкусВилл</div>
                    <h1>Отчёт по сверкам</h1>
                    <div class="subtitle">Админ-панель</div>
                </div>
                <div class="admin-export-buttons">
                    <a class="btn btn-secondary btn-inline" href="/admin-data">Управление данными</a>
                    <a class="btn btn-secondary btn-inline" href="/admin-logout">Выйти</a>
                </div>
            </div>

            <form method="get" action="/admin-report">
                <div class="filter-grid">
                    <div>
                        <label for="year">Год</label>
                        <input id="year" name="year" type="number" value="{year}" />
                    </div>

                    <div>
                        <label for="month">Месяц</label>
                        <select id="month" name="month">
                            {month_options}
                        </select>
                    </div>

                    <div>
                        <label for="tu">Территориальный управляющий</label>
                        <select id="tu" name="tu">
                            {tu_options}
                        </select>
                    </div>

                    <div>
                        <label for="status">Статус сверки</label>
                        <select id="status" name="status">
                            {status_options}
                        </select>
                    </div>
                </div>

                <button class="btn btn-inline" type="submit">Применить фильтр</button>
            </form>

            <div class="admin-export-buttons">
                <a class="btn btn-secondary btn-inline" href="/admin-export-check?{export_query}">Выгрузка для проверки</a>
                <a class="btn btn-secondary btn-inline" href="/admin-export-payroll?{export_query}">Выгрузка в ведомость</a>
                <a class="btn btn-secondary btn-inline" href="/admin-export-overlaps?{export_query}">Выгрузка пересечений</a>
            </div>

            <div class="hint">
                Период отчёта: {month_title(year, month)}. Всего строк: {len(rows)}.
            </div>

            <div class="table-wrap" style="margin-top:16px;">
                <table>
                    <thead>
                        <tr>
                            <th>ФИО</th>
                            <th>ТУ</th>
                            <th>Точка</th>
                            <th>Месяц</th>
                            <th>С поставкой</th>
                            <th>Без поставки</th>
                            <th>Инвенты</th>
                            <th>Кофемашина</th>
                            <th>Примечание по точке</th>
                            <th>Возмещение по точке</th>
                            <th>Итог по точке</th>
                            <th>Статус</th>
                            <th>Чек по точке</th>
                            <th>Комментарий месяца</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows_html}
                    </tbody>
                </table>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.get("/admin-data", response_class=HTMLResponse)
def admin_data_page(
    success: str = "",
    error: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    period = get_active_period()
    tu_values = get_all_tu_values(db)

    tu_options = ""
    for item in tu_values:
        tu_options += f"<option value='{escape(item)}'>{escape(item)}</option>"

    info_box = ""
    if success:
        info_box += f"<div class='success-box'>{escape(success)}</div>"
    if error:
        info_box += f"<div class='error-box'>{escape(error)}</div>"

    return f"""
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>Управление данными</title>
    {base_css()}
</head>
<body>
    <div class="page">
        <div class="card-wide">
            <div class="admin-actions">
                <div>
                    <div class="brand">ВкусВилл</div>
                    <h1>Управление данными</h1>
                    <div class="subtitle">Загрузка файлов и очистка месяца</div>
                </div>
                <div class="admin-export-buttons">
                    <a class="btn btn-secondary btn-inline" href="/admin-report">Назад к отчёту</a>
                    <a class="btn btn-secondary btn-inline" href="/admin-logout">Выйти</a>
                </div>
            </div>

            {info_box}

            <div class="data-grid">
                <div class="detail-card">
                    <div class="detail-title">Загрузка поставок</div>
                    <form method="post" action="/admin-upload-supplies" enctype="multipart/form-data">
                        <label for="supplies_file">Файл поставок</label>
                        <input id="supplies_file" name="file" type="file" accept=".xlsx" required />
                        <button class="btn" type="submit">Загрузить поставки</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Загрузка ставок</div>
                    <form method="post" action="/admin-upload-rates" enctype="multipart/form-data">
                        <label for="rates_year">Год</label>
                        <input id="rates_year" name="year" type="number" value="{period["year"]}" required />

                        <label for="rates_month">Месяц</label>
                        <input id="rates_month" name="month" type="number" value="{period["month"]}" min="1" max="12" required />

                        <label for="rates_file">Файл ставок</label>
                        <input id="rates_file" name="file" type="file" accept=".xlsx" required />

                        <button class="btn" type="submit">Загрузить ставки</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Загрузка мерчей</div>
                    <form method="post" action="/admin-upload-merchants" enctype="multipart/form-data">
                        <label for="merchants_tu">Территориальный управляющий</label>
                        <input id="merchants_tu" name="tu" type="text" placeholder="Например: Хрупов" required />

                        <label for="merchants_file">Файл мерчей</label>
                        <input id="merchants_file" name="file" type="file" accept=".xlsx" required />

                        <button class="btn" type="submit">Загрузить мерчей</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Очистка месяца</div>
                    <form method="post" action="/admin-clear-month">
                        <label for="clear_year">Год</label>
                        <input id="clear_year" name="year" type="number" value="{period["year"]}" required />

                        <label for="clear_month">Месяц</label>
                        <input id="clear_month" name="month" type="number" value="{period["month"]}" min="1" max="12" required />

                        <button class="btn btn-danger" type="submit">Очистить данные месяца</button>
                    </form>
                </div>

                <div class="detail-card">
                    <div class="detail-title">Очистка мерчей по ТУ</div>
                    <form method="post" action="/admin-clear-merchants">
                        <label for="clear_tu">Территориальный управляющий</label>
                        <select id="clear_tu" name="tu" required>
                            {tu_options}
                        </select>

                        <button class="btn btn-danger" type="submit">Удалить мерчей этого ТУ</button>
                    </form>
                </div>
            </div>
        </div>
    </div>
</body>
</html>
"""


@app.post("/admin-upload-supplies")
async def admin_upload_supplies(
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_supplies_xlsx(db, file.file)
        msg = f"Поставки загружены: строк {result['loaded_rows']}, точек {result['loaded_points']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-upload-rates")
async def admin_upload_rates(
    year: int = Form(...),
    month: int = Form(...),
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_rates_xlsx(db, file.file, year, month)
        msg = f"Ставки загружены: строк {result['loaded_rows']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-upload-merchants")
async def admin_upload_merchants(
    tu: str = Form(...),
    file: UploadFile = File(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    try:
        result = import_merchants_xlsx(db, file.file, tu)
        msg = f"Мерчи загружены: строк {result['loaded_rows']}."
        return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin-data?error={str(e)}", status_code=303)


@app.post("/admin-clear-month")
def admin_clear_month(
    year: int = Form(...),
    month: int = Form(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    result = clear_month_data(db, year, month)
    msg = (
        f"Месяц очищен. Визиты: {result['deleted_visits']}, "
        f"поставки: {result['deleted_supplies']}, ставки: {result['deleted_rates']}, "
        f"месячные сверки: {result['deleted_monthly']}, корректировки по точкам: {result.get('deleted_point_adjustments', 0)}."
    )
    return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)


@app.post("/admin-clear-merchants")
def admin_clear_merchants(
    tu: str = Form(...),
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    deleted = clear_merchants_by_tu(db, tu)
    msg = f"Удалено мерчей ТУ {tu}: {deleted}."
    return RedirectResponse(url=f"/admin-data?success={msg}", status_code=303)


@app.get("/admin-export-check")
def admin_export_check(
    year: int,
    month: int,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_admin_report_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None,
        status=status.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка"

    ws.append([
        "ФИО",
        "ТУ",
        "Точка",
        "Месяц",
        "Выходы с поставкой (кол-во)",
        "Выходы с поставкой (сумма)",
        "Выходы без поставки (кол-во)",
        "Выходы без поставки (сумма)",
        "Полные инвенты (кол-во)",
        "Полные инвенты (сумма)",
        "Кофемашина (кол-во)",
        "Кофемашина (сумма)",
        "Примечание по точке (сумма)",
        "Примечание по точке (комментарий)",
        "Возмещение по точке (сумма)",
        "Возмещение по точке (комментарий)",
        "Чек по точке",
        "Статус",
        "Комментарий месяца",
        "Итог по точке"
    ])

    for r in rows:
        ws.append([
            r["fio"],
            r["tu"],
            r["point_code"],
            month_title(year, month),
            r["cnt_supply"],
            r["sum_supply"],
            r["cnt_no_supply"],
            r["sum_no_supply"],
            r["cnt_full_inv"],
            r["sum_inventory"],
            r["coffee_cnt"],
            r["coffee_sum"],
            r["note_amount"],
            r["note_comment"],
            r["reimb_amount"],
            r["reimb_comment"],
            r["reimb_receipt"] or "",
            r["status"],
            r["comment"],
            r["point_total"]
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"proverka_{year}_{month:02d}.xlsx")


@app.get("/admin-export-payroll")
def admin_export_payroll(
    year: int,
    month: int,
    tu: str = "",
    status: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_admin_payroll_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None,
        status=status.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    ws.append([
        "ФИО",
        "ТУ",
        "Сумма по мерчу",
        "Сумма в ведомость (/0.87, округление вверх)",
        "Статус"
    ])

    for r in rows:
        ws.append([
            r["fio"],
            r["tu"],
            r["clean_total"],
            r["payroll_total"],
            r["status"]
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"vedomost_{year}_{month:02d}.xlsx")


@app.get("/admin-export-overlaps")
def admin_export_overlaps(
    year: int,
    month: int,
    tu: str = "",
    admin_auth: Optional[str] = Cookie(default=None),
    db: Session = Depends(get_db)
):
    if not is_admin_authenticated(admin_auth):
        return RedirectResponse(url="/admin-login", status_code=303)

    rows = get_intersections_rows(
        db=db,
        y=year,
        m=month,
        tu=tu.strip() or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Пересечения"

    ws.append([
        "Дата",
        "Точка",
        "Мерч 1",
        "ТУ 1",
        "Слот 1",
        "Мерч 2",
        "ТУ 2",
        "Слот 2"
    ])

    for r in rows:
        ws.append([
            r["visit_date"],
            r["point_code"],
            r["fio1"],
            r["tu1"],
            r["slot1"],
            r["fio2"],
            r["tu2"],
            r["slot2"],
        ])

    style_sheet(ws)
    return build_excel_response(wb, f"peresecheniya_{year}_{month:02d}.xlsx")

